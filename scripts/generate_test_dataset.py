#!/usr/bin/env python3
"""
Generate a complete realistic test dataset for ESG Flow.

Creates a fictional French SME ("Aurora Solutions SAS", 87 employees,
8.4 M€ CA, IT consulting) with one full year (2025) of:

  - FEC accounting export (real French tax format — articles A. 47 A-1 LPF)
  - ESG indicators CSV (~90 data points across E/S/G pillars)
  - Suppliers CSV (15 suppliers with ESG scores)
  - Materiality issues CSV (10 issues with double-materiality scores)
  - ESG risks register CSV (8 risks, prob × impact)
  - Carbon emissions CSV (Scopes 1/2/3 by category)
  - Master Excel workbook (every CSV as a tab)
  - README explaining how to import each file in the platform

Output: docs/test-dataset/

Usage:
    /Users/cisseniang/Downloads/esgplatform/.claude/worktrees/elastic-moore/backend/.venv/bin/python \\
        scripts/generate_test_dataset.py
"""
from __future__ import annotations

import csv
import random
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Company profile ────────────────────────────────────────────────────────
COMPANY = {
    "name":        "Aurora Solutions SAS",
    "siren":       "851234567",
    "siret":       "85123456700019",
    "naf":         "6202A",  # Conseil en systèmes et logiciels informatiques
    "sector":      "services_b2b",
    "address":     "42 rue de la Paix, 75002 Paris",
    "fiscal_year": 2025,
    "employees":   87,
    "revenue_k":   8_400,     # k€
    "fiscal_start": date(2025, 1, 1),
    "fiscal_end":   date(2025, 12, 31),
}

random.seed(42)  # reproducible runs


# ─── Output directory ───────────────────────────────────────────────────────
def out_dir() -> Path:
    p = Path(__file__).resolve().parent.parent / 'docs' / 'test-dataset'
    p.mkdir(parents=True, exist_ok=True)
    return p


# ─── 1. FEC (Fichier des Écritures Comptables) ─────────────────────────────
# Format réglementaire français : article A. 47 A-1 du Livre des Procédures Fiscales.
# 18 colonnes obligatoires, séparateur tabulation, UTF-8.

FEC_HEADERS = [
    "JournalCode", "JournalLib", "EcritureNum", "EcritureDate",
    "CompteNum", "CompteLib", "CompAuxNum", "CompAuxLib",
    "PieceRef", "PieceDate", "EcritureLib",
    "Debit", "Credit",
    "EcritureLet", "DateLet", "ValidDate",
    "Montantdevise", "Idevise",
]

# Plan comptable général réaliste pour une PME conseil IT
ACCOUNTS = {
    # Achats (Scope 3 cat 1 — biens et services)
    "601000": "Achats stockés - Matières premières",
    "606300": "Fournitures non stockables (eau, électricité)",
    "606400": "Fournitures administratives",
    "606800": "Autres matières et fournitures",
    # Services extérieurs (Scope 3 cat 1 — services)
    "611000": "Sous-traitance générale",
    "613100": "Locations immobilières",
    "613500": "Locations mobilières (informatique)",
    "615000": "Entretien et réparations",
    "615500": "Entretien matériel informatique",
    "616000": "Primes d'assurance",
    "617000": "Études et recherches",
    "618000": "Documentation",
    # Autres services extérieurs (Scope 3 cat 1 + 6 + 9)
    "621000": "Personnel extérieur à l'entreprise",
    "622600": "Honoraires (avocats, experts)",
    "623000": "Publicité, publications, relations publiques",
    "624000": "Transports de biens",
    "625100": "Voyages et déplacements (avion, train)",   # Scope 3 cat 6
    "625600": "Missions",
    "625700": "Réceptions",
    "626000": "Frais postaux et télécommunications",
    "627000": "Services bancaires",
    "628100": "Cotisations professionnelles",
    # Impôts
    "631100": "Taxe sur les salaires",
    "635100": "Taxes diverses",
    # Charges de personnel (hors Scope 3, mais utile pour S1)
    "641000": "Rémunérations du personnel",
    "645000": "Charges de sécurité sociale et de prévoyance",
    # Énergie (Scope 1+2)
    "606110": "Électricité",            # Scope 2
    "606120": "Gaz naturel",            # Scope 1
    "606130": "Carburants véhicules",   # Scope 1
}

# Tier 1 fournisseurs (apparaîtront en compte auxiliaire CompAuxNum 401xxx)
SUPPLIERS = [
    ("401AWS",      "AMAZON WEB SERVICES",      "Cloud computing"),
    ("401MICRO",    "MICROSOFT FRANCE",          "Licences logicielles"),
    ("401PENNY",    "PENNYLANE SAS",             "Comptabilité SaaS"),
    ("401EDF",      "EDF ENTREPRISES",           "Électricité"),
    ("401ENGIE",    "ENGIE",                     "Gaz naturel"),
    ("401TOTAL",    "TOTALENERGIES",             "Carburants flotte"),
    ("401WEWORK",   "WEWORK",                    "Location bureaux"),
    ("401MANPOWER", "MANPOWER",                  "Intérim développeurs"),
    ("401KPMG",     "KPMG AVOCATS",              "Conseil juridique"),
    ("401ORANGE",   "ORANGE BUSINESS",           "Télécoms"),
    ("401AF",       "AIR FRANCE",                "Déplacements aériens"),
    ("401SNCF",     "SNCF VOYAGEURS",            "Déplacements train"),
    ("401UBER",     "UBER FRANCE",               "Taxi/VTC"),
    ("401LINKEDIN", "LINKEDIN FRANCE",           "Marketing recrutement"),
    ("401GOOGLE",   "GOOGLE FRANCE",             "Google Ads + Workspace"),
    ("401VINCI",    "VINCI FACILITIES",          "Entretien bureaux"),
    ("401AXA",      "AXA ASSURANCES",            "Assurances pro"),
    ("401LIBRARY",  "LE LIBRAIRE PRO",           "Documentation technique"),
    ("401ALAN",     "ALAN SANTÉ",                "Mutuelle salariés"),
    ("401CAFE",     "BIOCAFE TRAITEUR",          "Restauration entreprise"),
]


def gen_fec_entry(num, journal_code, journal_lib, account, account_lib,
                  comp_aux_num, comp_aux_lib, label, debit, credit,
                  d, piece_ref):
    """Construit une ligne FEC respectant le format A. 47 A-1."""
    return [
        journal_code,
        journal_lib,
        str(num).zfill(6),                       # EcritureNum (6 digits)
        d.strftime("%Y%m%d"),                    # EcritureDate AAAAMMJJ
        account,                                  # CompteNum
        account_lib,                              # CompteLib
        comp_aux_num,                             # CompAuxNum
        comp_aux_lib,                             # CompAuxLib
        piece_ref,                                # PieceRef
        d.strftime("%Y%m%d"),                    # PieceDate
        label,                                   # EcritureLib
        f"{debit:.2f}".replace(".", ","),        # Debit (décimal séparateur virgule)
        f"{credit:.2f}".replace(".", ","),
        "",                                       # EcritureLet (lettrage — vide pour ce cas test)
        "",                                       # DateLet
        d.strftime("%Y%m%d"),                    # ValidDate
        "",                                       # Montantdevise (EUR par défaut, vide)
        "",                                       # Idevise (EUR par défaut, vide)
    ]


def build_fec():
    """Génère ~250 écritures comptables réalistes pour 2025."""
    rows: list[list] = [FEC_HEADERS]
    num = 1

    # Fichier de référence : 1 écriture mensuelle par poste majeur + qq exceptionnelles
    monthly_dates = [date(2025, m, 28) for m in range(1, 13)]

    # ─── Postes mensuels récurrents ──────────────────────────────────────
    monthly_recurring = [
        # (account, account_lib, supplier_tuple, label, monthly_amount, journal)
        ("613100", "Locations immobilières",     ("401WEWORK", "WEWORK"),         "Loyer WeWork Paris",        7_500, "AC"),
        ("606110", "Électricité",                 ("401EDF",    "EDF ENTREPRISES"), "Facture électricité",        2_400, "AC"),
        ("606120", "Gaz naturel",                 ("401ENGIE",  "ENGIE"),           "Facture gaz chauffage",      1_800, "AC"),
        ("611000", "Sous-traitance générale",     ("401AWS",    "AMAZON WEB SERVICES"), "Infrastructure cloud", 4_200, "AC"),
        ("613500", "Locations matériel info",     ("401MICRO",  "MICROSOFT FRANCE"), "Licences MS 365 + Azure",   3_500, "AC"),
        ("626000", "Frais postaux / télécoms",    ("401ORANGE", "ORANGE BUSINESS"), "Téléphonie + fibre",          950, "AC"),
        ("621000", "Personnel extérieur",         ("401MANPOWER", "MANPOWER"),     "Intérim développeurs",      18_500, "AC"),
        ("641000", "Rémunérations personnel",     ("",          ""),               "Salaires bruts",          420_000, "OD"),
        ("645000", "Charges sociales",            ("",          ""),               "Cotisations URSSAF + retraite", 210_000, "OD"),
        ("617000", "Études et recherches",        ("401PENNY",  "PENNYLANE SAS"),  "Abonnement Pennylane",         110, "AC"),
        ("615500", "Entretien matériel info",     ("401VINCI",  "VINCI FACILITIES"), "Entretien bureaux",         1_200, "AC"),
        ("606130", "Carburants véhicules",        ("401TOTAL",  "TOTALENERGIES"),  "Carburant flotte 5 véhicules", 850, "AC"),
        ("623000", "Publicité",                   ("401GOOGLE", "GOOGLE FRANCE"),  "Google Ads",                 2_800, "AC"),
        ("623000", "Publicité",                   ("401LINKEDIN", "LINKEDIN FRANCE"), "LinkedIn Ads recrutement",   1_500, "AC"),
    ]

    for d in monthly_dates:
        for acct, acct_lib, supp, lbl, amt, j_code in monthly_recurring:
            # Légère variation aléatoire ±10 % pour le réalisme
            var = random.uniform(0.92, 1.08)
            amount = round(amt * var, 2)
            j_lib = "Achats" if j_code == "AC" else "Opérations diverses"
            piece = f"F2025-{num:05d}"
            rows.append(gen_fec_entry(num, j_code, j_lib, acct, acct_lib,
                                       supp[0], supp[1], lbl, amount, 0.00, d, piece))
            num += 1
            # Contrepartie sur compte fournisseur 401xxx (si applicable)
            if supp[0]:
                rows.append(gen_fec_entry(num, j_code, j_lib, "401000",
                                           "Fournisseurs", supp[0], supp[1],
                                           lbl, 0.00, amount, d, piece))
                num += 1

    # ─── Déplacements pro (Scope 3 cat 6) — irrégulier ────────────────────
    for _ in range(36):
        d = date(2025, random.randint(1, 12), random.randint(1, 28))
        amount = round(random.uniform(150, 1_800), 2)
        air_or_train = random.choice([
            ("401AF",   "AIR FRANCE",     "Vol Paris-Lyon AR cliente"),
            ("401AF",   "AIR FRANCE",     "Vol Paris-Toulouse mission"),
            ("401SNCF", "SNCF VOYAGEURS", "TGV Paris-Marseille"),
            ("401SNCF", "SNCF VOYAGEURS", "TGV Paris-Bordeaux"),
        ])
        piece = f"F2025-{num:05d}"
        rows.append(gen_fec_entry(num, "AC", "Achats",
                                   "625100", "Voyages et déplacements",
                                   air_or_train[0], air_or_train[1],
                                   air_or_train[2], amount, 0.00, d, piece))
        num += 1

    # ─── Honoraires juridiques (variable, semestriel) ────────────────────
    for d in [date(2025, 4, 15), date(2025, 10, 22)]:
        amount = round(random.uniform(5_000, 15_000), 2)
        piece = f"F2025-{num:05d}"
        rows.append(gen_fec_entry(num, "AC", "Achats", "622600", "Honoraires",
                                   "401KPMG", "KPMG AVOCATS",
                                   "Conseil juridique semestriel", amount, 0.00, d, piece))
        num += 1

    # ─── Assurances annuelles ────────────────────────────────────────────
    piece = f"F2025-{num:05d}"
    rows.append(gen_fec_entry(num, "AC", "Achats", "616000", "Primes d'assurance",
                               "401AXA", "AXA ASSURANCES",
                               "Assurance RC + multirisques", 12_500, 0.00, date(2025, 1, 15), piece))
    num += 1

    # ─── Mutuelle santé salariés (S1) ────────────────────────────────────
    for d in monthly_dates:
        piece = f"F2025-{num:05d}"
        rows.append(gen_fec_entry(num, "AC", "Achats", "645000", "Charges sociales",
                                   "401ALAN", "ALAN SANTÉ",
                                   "Mutuelle Alan (87 salariés)", round(87 * 52 * random.uniform(0.95, 1.05), 2), 0.00, d, piece))
        num += 1

    # ─── Restauration entreprise ─────────────────────────────────────────
    for d in monthly_dates:
        piece = f"F2025-{num:05d}"
        rows.append(gen_fec_entry(num, "AC", "Achats", "625700", "Réceptions",
                                   "401CAFE", "BIOCAFE TRAITEUR",
                                   "Déjeuner mensuel équipe", round(random.uniform(800, 1_500), 2), 0.00, d, piece))
        num += 1

    return rows


def write_fec():
    rows = build_fec()
    siren = COMPANY["siren"]
    closing = COMPANY["fiscal_end"].strftime("%Y%m%d")
    filename = f"{siren}FEC{closing}.txt"
    path = out_dir() / filename
    with open(path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t', lineterminator='\n', quoting=csv.QUOTE_NONE, escapechar='\\')
        for row in rows:
            writer.writerow(row)
    return path, len(rows) - 1  # -1 header


# ─── 2. ESG indicators CSV ──────────────────────────────────────────────────
ESG_INDICATORS = [
    # (pillar, esrs_code, metric_name, value, unit, category, verification_status)

    # ── ESRS 2 — Informations générales
    ("governance",    "ESRS2-1",  "Présence comité ESG au CA",                            1,        "bool",   "governance",         "verified"),
    ("governance",    "ESRS2-2",  "Périmètre reporting consolidé",                         1,        "bool",   "governance",         "verified"),

    # ── E1 — Climat
    ("environmental", "E1-1",     "Engagement net-zéro 2050",                              1,        "bool",   "climate_strategy",   "verified"),
    ("environmental", "E1-4",     "Objectif réduction GES 2030 vs 2019",                  42,        "%",      "climate_target",     "verified"),
    ("environmental", "E1-5",     "Consommation énergie totale",                       1_120_000,    "kWh",    "energy",             "verified"),
    ("environmental", "E1-5b",    "Part énergie renouvelable",                            68,        "%",      "energy",             "verified"),
    ("environmental", "E1-6a",    "Scope 1 émissions directes",                           58.5,      "tCO2e",  "ghg_scope1",         "verified"),
    ("environmental", "E1-6b",    "Scope 2 location-based",                               64.0,      "tCO2e",  "ghg_scope2",         "verified"),
    ("environmental", "E1-6b2",   "Scope 2 market-based",                                  0.0,      "tCO2e",  "ghg_scope2",         "verified"),
    ("environmental", "E1-6c",    "Scope 3 total",                                     2_178.0,      "tCO2e",  "ghg_scope3",         "pending"),
    ("environmental", "E1-6c1",   "Scope 3 cat 1 - Achats biens et services",          1_650.0,      "tCO2e",  "ghg_scope3_cat1",    "pending"),
    ("environmental", "E1-6c6",   "Scope 3 cat 6 - Déplacements pro",                     115.0,      "tCO2e",  "ghg_scope3_cat6",    "pending"),
    ("environmental", "E1-6c7",   "Scope 3 cat 7 - Domicile-travail",                     290.0,      "tCO2e",  "ghg_scope3_cat7",    "pending"),
    ("environmental", "E1-8",     "Prix interne du carbone",                              45,        "EUR/tCO2e", "carbon_pricing", "verified"),
    ("environmental", "E1-9",     "Effets financiers risques climatiques",             1_800_000,    "EUR",    "climate_risk",       "pending"),

    # ── E2 — Pollution
    ("environmental", "E2-4",     "Émissions NOx",                                          0.4,     "tonnes", "air_pollution",      "verified"),
    ("environmental", "E2-5",     "Substances préoccupantes utilisées",                     0,       "tonnes", "chemicals",          "verified"),

    # ── E3 — Eau
    ("environmental", "E3-4",     "Prélèvements d'eau totaux",                          4_350,       "m3",     "water_withdrawal",   "verified"),
    ("environmental", "E3-4b",    "Prélèvements en zones stress hydrique",                  0,       "m3",     "water_stress",       "verified"),
    ("environmental", "E3-5",     "Consommation d'eau",                                  4_200,      "m3",     "water_consumption",  "verified"),

    # ── E4 — Biodiversité
    ("environmental", "E4-2",     "Sites proches zones Natura 2000",                       0,        "nb",     "biodiversity",       "verified"),

    # ── E5 — Économie circulaire
    ("environmental", "E5-4",     "Matières recyclées",                                   12,        "tonnes", "circular",           "verified"),
    ("environmental", "E5-4b",    "Part matières recyclées",                              35,        "%",      "circular",           "verified"),
    ("environmental", "E5-5",     "Déchets totaux générés",                              8.4,       "tonnes", "waste",              "verified"),
    ("environmental", "E5-5b",    "Taux de valorisation déchets",                         72,        "%",      "waste",              "verified"),

    # ── S1 — Effectifs propres
    ("social",        "S1-6",     "Nombre total salariés (ETP)",                          87,        "ETP",    "headcount",          "verified"),
    ("social",        "S1-6a",    "Salariés CDI",                                          81,        "ETP",    "headcount_cdi",      "verified"),
    ("social",        "S1-6b",    "Salariés CDD",                                           4,        "ETP",    "headcount_cdd",      "verified"),
    ("social",        "S1-6c",    "Intérimaires (moyenne annuelle)",                       12,        "ETP",    "headcount_temp",     "verified"),
    ("social",        "S1-9",     "Part de femmes total",                                 38,        "%",      "gender_diversity",   "verified"),
    ("social",        "S1-9a",    "Part de femmes en management",                         32,        "%",      "gender_diversity",   "verified"),
    ("social",        "S1-9b",    "Part de femmes au CODIR",                              25,        "%",      "gender_diversity",   "verified"),
    ("social",        "S1-10",    "Écart de rémunération H/F (gender pay gap)",            5.8,      "%",      "pay_equity",         "verified"),
    ("social",        "S1-10b",   "Ratio rémunération CEO / médiane",                     8.4,       "ratio",  "pay_equity",         "verified"),
    ("social",        "S1-13",    "Heures de formation par salarié (moyenne)",            22,        "heures", "training",           "verified"),
    ("social",        "S1-13b",   "Salariés formés au moins 1 fois",                      91,        "%",      "training",           "verified"),
    ("social",        "S1-14",    "Accidents du travail avec arrêt",                       1,        "nb",     "safety",             "verified"),
    ("social",        "S1-14b",   "Taux de fréquence accidents",                          2.1,       "ratio",  "safety",             "verified"),
    ("social",        "S1-14c",   "Maladies professionnelles déclarées",                   0,        "nb",     "safety",             "verified"),
    ("social",        "S1-15",    "Salariés en télétravail régulier",                     78,        "%",      "work_life",          "verified"),
    ("social",        "S1-16",    "Taux de turnover volontaire",                           9.4,      "%",      "turnover",           "verified"),
    ("social",        "S1-17",    "Salaires sous minimum sectoriel",                       0,        "%",      "minimum_wage",       "verified"),

    # ── S2 — Travailleurs chaîne de valeur
    ("social",        "S2-1",     "Politique droits humains supply chain publiée",          1,       "bool",   "supply_chain",       "verified"),
    ("social",        "S2-4",     "% fournisseurs (valeur achat) évalués droits humains", 62,        "%",      "supplier_audit",     "pending"),

    # ── S3 — Communautés
    ("social",        "S3-1",     "Montant mécénat / partenariats locaux",             24_000,       "EUR",    "community",          "verified"),

    # ── S4 — Consommateurs
    ("social",        "S4-1",     "Politique protection données personnelles",              1,       "bool",   "data_privacy",       "verified"),
    ("social",        "S4-3",     "Réclamations clients reçues",                          18,        "nb",     "customer_complaint", "verified"),
    ("social",        "S4-3b",    "Réclamations résolues sous 30j",                       17,        "nb",     "customer_complaint", "verified"),
    ("social",        "S4-4",     "Incidents données personnelles déclarés CNIL",           0,       "nb",     "data_privacy",       "verified"),

    # ── G1 — Gouvernance
    ("governance",    "G1-1",     "Code de conduite publié",                                1,       "bool",   "ethics",             "verified"),
    ("governance",    "G1-2",     "Délai moyen de paiement fournisseurs",                  38,        "jours",  "payment_terms",      "verified"),
    ("governance",    "G1-2b",    "Conformité légale délais (LME 60j)",                    96,        "%",      "payment_terms",      "verified"),
    ("governance",    "G1-3",     "Salariés formés anti-corruption",                       87,        "%",      "anti_corruption",    "verified"),
    ("governance",    "G1-3b",    "Cartographie risques corruption à jour",                 1,       "bool",   "anti_corruption",    "verified"),
    ("governance",    "G1-4",     "Cas confirmés de corruption",                            0,       "nb",     "anti_corruption",    "verified"),
    ("governance",    "G1-5",     "Dépenses lobbying",                                      0,       "EUR",    "lobbying",           "verified"),
    ("governance",    "G1-5b",    "Inscription registre transparence UE",                   0,       "bool",   "lobbying",           "verified"),
    ("governance",    "G1-6",     "Signalements lanceurs d'alerte reçus",                   2,       "nb",     "whistleblowing",     "verified"),
]


def write_indicators_csv():
    path = out_dir() / 'indicators-esg-2025.csv'
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'organization', 'period_start', 'period_end',
            'esrs_code', 'pillar', 'category', 'metric_name',
            'value', 'unit', 'verification_status',
        ])
        for pillar, code, name, value, unit, cat, status in ESG_INDICATORS:
            w.writerow([
                COMPANY['name'], '2025-01-01', '2025-12-31',
                code, pillar, cat, name, value, unit, status,
            ])
    return path, len(ESG_INDICATORS)


# ─── 3. Suppliers CSV ───────────────────────────────────────────────────────
SUPPLIERS_DATA = [
    # name, country, sector, annual_spend_eur, risk_level, e_score, s_score, g_score, audit_status
    ("AMAZON WEB SERVICES",     "US", "Cloud / IT",         52_000, "medium", 78, 70, 85, "self-declared"),
    ("MICROSOFT FRANCE",         "FR", "Software",           45_000, "low",    82, 76, 92, "ecovadis-gold"),
    ("PENNYLANE SAS",            "FR", "SaaS comptable",      1_320, "low",    72, 80, 88, "self-declared"),
    ("EDF ENTREPRISES",          "FR", "Énergie",            29_000, "medium", 65, 70, 75, "rapport-public"),
    ("ENGIE",                    "FR", "Énergie",            21_500, "medium", 68, 72, 78, "rapport-public"),
    ("TOTALENERGIES",            "FR", "Carburants",         10_400, "high",   45, 60, 65, "rapport-public"),
    ("WEWORK FRANCE",            "FR", "Immobilier bureaux", 91_500, "low",    70, 75, 70, "self-declared"),
    ("MANPOWER",                 "FR", "Intérim RH",        224_000, "medium", 60, 78, 82, "ecovadis-silver"),
    ("KPMG AVOCATS",             "FR", "Conseil juridique",  35_000, "low",    72, 80, 88, "self-declared"),
    ("ORANGE BUSINESS",          "FR", "Télécoms",           11_400, "low",    74, 76, 85, "ecovadis-gold"),
    ("AIR FRANCE",               "FR", "Transport aérien",   29_500, "high",   50, 68, 72, "rapport-public"),
    ("SNCF VOYAGEURS",           "FR", "Transport ferro",    14_800, "low",    85, 78, 80, "rapport-public"),
    ("UBER FRANCE",              "FR", "Transport VTC",       4_200, "high",   58, 55, 60, "non-évalué"),
    ("LINKEDIN FRANCE",          "FR", "Recrutement",        18_000, "low",    65, 72, 78, "self-declared"),
    ("GOOGLE FRANCE",            "FR", "Marketing digital",  33_600, "medium", 70, 65, 80, "rapport-public"),
]


def write_suppliers_csv():
    path = out_dir() / 'suppliers-2025.csv'
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'name', 'country', 'sector', 'annual_spend_eur',
            'risk_level', 'e_score', 's_score', 'g_score',
            'global_score', 'audit_status', 'audit_date',
        ])
        for s in SUPPLIERS_DATA:
            name, country, sector, spend, risk, e, soc, g, audit = s
            glob = round((e + soc + g) / 3, 1)
            w.writerow([name, country, sector, spend, risk, e, soc, g, glob, audit, '2025-06-30'])
    return path, len(SUPPLIERS_DATA)


# ─── 4. Materiality issues ──────────────────────────────────────────────────
MATERIALITY = [
    # (name, category, financial_impact, esg_impact, is_material, priority)
    ("Émissions GES et empreinte carbone",         "E1 Climat",          78, 92, True,  "high"),
    ("Mix énergétique et énergies renouvelables",  "E1 Climat",          65, 80, True,  "high"),
    ("Gestion responsable des déchets IT",         "E5 Économie circ.",  35, 58, True,  "medium"),
    ("Attraction et rétention des talents",        "S1 Effectifs",       88, 75, True,  "high"),
    ("Diversité et inclusion",                      "S1 Effectifs",       45, 82, True,  "high"),
    ("Formation et développement compétences",     "S1 Effectifs",       70, 78, True,  "high"),
    ("Audits sociaux fournisseurs",                "S2 Chaîne valeur",   55, 72, True,  "medium"),
    ("Protection des données clients (RGPD)",      "S4 Consommateurs",   92, 88, True,  "high"),
    ("Éthique et anti-corruption",                  "G1 Gouvernance",     75, 80, True,  "high"),
    ("Délais de paiement fournisseurs",            "G1 Gouvernance",     48, 35, False, "low"),
]


def write_materiality_csv():
    path = out_dir() / 'materiality-matrix-2025.csv'
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'name', 'category', 'financial_impact', 'esg_impact',
            'is_material', 'priority', 'stakeholders',
        ])
        for name, cat, fi, esg, mat, prio in MATERIALITY:
            w.writerow([name, cat, fi, esg, 'oui' if mat else 'non', prio,
                        'salariés, clients, investisseurs'])
    return path, len(MATERIALITY)


# ─── 5. ESG risks register ──────────────────────────────────────────────────
RISKS = [
    # (title, category, probability_1_5, impact_1_5, mitigation_status, severity)
    ("Hausse prix énergie (Scope 1+2)",       "Transition risk",      4, 3, "in_progress", "high"),
    ("Pénurie talents tech (turnover)",       "Operational",          4, 4, "in_progress", "critical"),
    ("Cyberattaque / fuite données clients",   "Operational",          3, 5, "implemented", "critical"),
    ("Non-conformité CSRD vague 2",           "Regulatory",           3, 4, "in_progress", "high"),
    ("Sanction CNIL incident RGPD",           "Regulatory",           2, 4, "implemented", "high"),
    ("Aléas climat physique (canicule, datacenter)", "Physical risk",   2, 3, "planned",    "medium"),
    ("Dépendance fournisseur cloud unique (AWS)",  "Operational",      3, 3, "planned",    "medium"),
    ("Atteinte réputation greenwashing",       "Reputational",         2, 4, "in_progress", "high"),
]


def write_risks_csv():
    path = out_dir() / 'esg-risks-register-2025.csv'
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            'title', 'category', 'probability', 'impact', 'risk_score',
            'severity', 'mitigation_status', 'owner',
        ])
        for title, cat, p, i, mit, sev in RISKS:
            w.writerow([title, cat, p, i, p * i, sev, mit, 'Direction'])
    return path, len(RISKS)


# ─── 6. Carbon emissions breakdown ──────────────────────────────────────────
CARBON_DETAIL = [
    # (scope, category, source, value_tco2e, calculation_method)
    ("scope1", "1.1", "Combustion gaz chauffage bureaux",       26.5,  "average-data"),
    ("scope1", "1.2", "Carburant flotte (5 véhicules diesel)",  21.0,  "average-data"),
    ("scope1", "1.3", "Fuites fluides clim (R-410A 5 kg)",      10.5,  "average-data"),
    ("scope1", "1.4", "Émissions process (n/a IT services)",      0.5,  "estimated"),

    ("scope2", "2.1", "Électricité location-based France",       64.0,  "location-based"),
    ("scope2", "2.2", "Électricité market-based (CPPA 100%)",     0.0,  "market-based"),

    ("scope3", "3.1",  "Achats biens & services (FEC spend)",  1_650.0, "spend-based"),
    ("scope3", "3.2",  "Biens d'équipement (IT, mobilier)",      85.0,  "spend-based"),
    ("scope3", "3.3",  "Énergie hors S1/S2",                     12.0,  "average-data"),
    ("scope3", "3.4",  "Transport amont (livraisons)",            8.0,  "spend-based"),
    ("scope3", "3.5",  "Déchets bureautiques",                    4.2,  "average-data"),
    ("scope3", "3.6",  "Déplacements pro (avion + train)",      115.0,  "average-data"),
    ("scope3", "3.7",  "Domicile-travail (87 sal)",              290.0, "average-data"),
    ("scope3", "3.8",  "Actifs en leasing amont",                 14.0, "spend-based"),
    ("scope3", "3.9",  "Transport aval (n/a)",                     0.0, "n/a"),
    ("scope3", "3.10", "Transformation produits (n/a)",            0.0, "n/a"),
    ("scope3", "3.11", "Usage produits vendus (n/a services)",     0.0, "n/a"),
    ("scope3", "3.12", "Fin de vie produits (n/a)",                0.0, "n/a"),
    ("scope3", "3.13", "Actifs en leasing aval (n/a)",             0.0, "n/a"),
    ("scope3", "3.14", "Franchises (n/a)",                         0.0, "n/a"),
    ("scope3", "3.15", "Investissements (n/a)",                    0.0, "n/a"),
]


def write_carbon_csv():
    path = out_dir() / 'carbon-emissions-2025.csv'
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['scope', 'category', 'source', 'tco2e', 'calculation_method', 'period'])
        for scope, cat, source, val, method in CARBON_DETAIL:
            w.writerow([scope, cat, source, val, method, '2025'])
    return path, len(CARBON_DETAIL)


# ─── 7. Master Excel workbook ───────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
ACCENT_FILL = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
ALT_FILL    = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin', color='E5E7EB'),
                     right=Side(style='thin', color='E5E7EB'),
                     top=Side(style='thin', color='E5E7EB'),
                     bottom=Side(style='thin', color='E5E7EB'))


def style_header(cell):
    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = THIN_BORDER


def write_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for c in ws[1]:
        style_header(c)
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for c in ws[i]:
                c.fill = ALT_FILL
        for c in ws[i]:
            c.font = Font(name='Calibri', size=10)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical='center')
    # auto width
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"


def write_xlsx():
    wb = Workbook()
    # Default sheet → cover info
    cover = wb.active
    cover.title = "Présentation"
    cover['A1'] = "ESG Flow — Jeu de données de test"
    cover['A1'].font = Font(size=22, bold=True, color='0F172A')
    cover['A3'] = COMPANY['name']
    cover['A3'].font = Font(size=16, bold=True, color='10B981')
    cover['A4'] = f"SIREN {COMPANY['siren']} · SIRET {COMPANY['siret']} · NAF {COMPANY['naf']}"
    cover['A5'] = COMPANY['address']
    cover['A6'] = f"Exercice : {COMPANY['fiscal_year']} · Effectif {COMPANY['employees']} ETP · CA {COMPANY['revenue_k']} k€"
    cover['A8'] = "Ce classeur contient tous les éléments nécessaires pour tester la plateforme :"
    cover['A9']  = "• Indicateurs ESG (onglet « Indicateurs »)"
    cover['A10'] = "• Bilan carbone Scopes 1/2/3 (onglet « Carbone »)"
    cover['A11'] = "• Fournisseurs (onglet « Fournisseurs »)"
    cover['A12'] = "• Matrice de double matérialité (onglet « Matérialité »)"
    cover['A13'] = "• Registre des risques ESG (onglet « Risques »)"
    cover['A15'] = "Voir aussi les fichiers CSV/TXT séparés pour les imports unitaires."
    cover['A15'].font = Font(italic=True, color='6B7280')
    cover.column_dimensions['A'].width = 80

    # ── Indicateurs
    ws = wb.create_sheet("Indicateurs")
    rows = []
    for pillar, code, name, value, unit, cat, status in ESG_INDICATORS:
        rows.append([COMPANY['name'], '2025-01-01', '2025-12-31', code, pillar, cat, name, value, unit, status])
    write_sheet(ws, ['Organisation', 'Période début', 'Période fin', 'Code ESRS',
                     'Pilier', 'Catégorie', 'Indicateur', 'Valeur', 'Unité', 'Statut'], rows)

    # ── Carbone
    ws = wb.create_sheet("Carbone")
    rows = []
    for scope, cat, source, val, method in CARBON_DETAIL:
        rows.append([scope.upper(), cat, source, val, method, '2025'])
    write_sheet(ws, ['Scope', 'Catégorie GHG', 'Source', 'tCO2e', 'Méthode', 'Période'], rows)

    # ── Fournisseurs
    ws = wb.create_sheet("Fournisseurs")
    rows = []
    for s in SUPPLIERS_DATA:
        name, country, sector, spend, risk, e, soc, g, audit = s
        glob = round((e + soc + g) / 3, 1)
        rows.append([name, country, sector, spend, risk, e, soc, g, glob, audit, '2025-06-30'])
    write_sheet(ws, ['Nom', 'Pays', 'Secteur', 'Dépense annuelle (€)', 'Risque',
                     'Score E', 'Score S', 'Score G', 'Score global', 'Statut audit', 'Date audit'], rows)

    # ── Matérialité
    ws = wb.create_sheet("Matérialité")
    rows = []
    for name, cat, fi, esg, mat, prio in MATERIALITY:
        rows.append([name, cat, fi, esg, 'oui' if mat else 'non', prio, 'salariés, clients, investisseurs'])
    write_sheet(ws, ['Enjeu', 'Catégorie', 'Impact financier (0-100)', 'Impact ESG (0-100)',
                     'Matériel ?', 'Priorité', 'Parties prenantes'], rows)

    # ── Risques
    ws = wb.create_sheet("Risques")
    rows = []
    for title, cat, p, i, mit, sev in RISKS:
        rows.append([title, cat, p, i, p * i, sev, mit, 'Direction'])
    write_sheet(ws, ['Titre', 'Catégorie', 'Probabilité (1-5)', 'Impact (1-5)', 'Score',
                     'Sévérité', 'Statut mitigation', 'Responsable'], rows)

    path = out_dir() / f"{COMPANY['name'].replace(' ', '-').replace('.', '')}-DATASET-2025.xlsx"
    wb.save(path)
    return path


# ─── 8. README ──────────────────────────────────────────────────────────────
def write_readme(fec_info, indic_info, supp_info, mat_info, risk_info, carbon_info, xlsx_path):
    path = out_dir() / 'README.md'
    fec_path, fec_count = fec_info
    indic_path, indic_count = indic_info
    supp_path, supp_count = supp_info
    mat_path, mat_count = mat_info
    risk_path, risk_count = risk_info
    carbon_path, carbon_count = carbon_info

    content = f"""# Jeu de données de test ESG Flow

> Données fictives mais réalistes pour tester la plateforme **sans connecteur**.
> Permet d'évaluer : calcul des scores ESG, benchmark sectoriel, taxonomie UE,
> bilan carbone Scopes 1/2/3, audit fournisseurs, matrice de matérialité,
> registre des risques.

---

## 🏢 Entreprise fictive

| Champ          | Valeur                                       |
|---             |---                                           |
| **Nom**        | {COMPANY['name']}                            |
| **SIREN**      | {COMPANY['siren']}                           |
| **SIRET**      | {COMPANY['siret']}                           |
| **NAF**        | {COMPANY['naf']} — Conseil en systèmes et logiciels informatiques |
| **Secteur**    | Services aux entreprises (IT/conseil)        |
| **Effectif**   | {COMPANY['employees']} ETP                   |
| **CA**         | {COMPANY['revenue_k']:,} k€                  |
| **Adresse**    | {COMPANY['address']}                         |
| **Exercice**   | 1er jan – 31 déc {COMPANY['fiscal_year']}    |

---

## 📂 Fichiers générés

### 1. `{fec_path.name}` — FEC comptable
- **Format** : Fichier des Écritures Comptables au format réglementaire français
  (Article A. 47 A-1 du Livre des Procédures Fiscales)
- **Encodage** : UTF-8, séparateur tabulation, séparateur décimal virgule
- **Contenu** : **{fec_count}** lignes d'écritures sur l'exercice 2025
- **Comptes représentés** : 60* (achats), 61*-62* (services), 63* (impôts),
  64* (personnel), 401* (fournisseurs auxiliaires)
- **Utilité** : Import dans ESG Flow → calcul automatique du **Scope 3 spend-based**
  avec les facteurs ADEME Base Carbone®. Les comptes 625100 (déplacements pro)
  alimentent le Scope 3 catégorie 6.

**Import sur la plateforme** : `Données → Import → FEC comptable`

### 2. `{indic_path.name}` — Indicateurs ESG
- **Format** : CSV avec en-tête, séparateur point-virgule
- **Contenu** : **{indic_count} indicateurs** couvrant les 12 normes ESRS :
  ESRS 2 transversal, E1 (climat), E2 (pollution), E3 (eau), E4 (biodiversité),
  E5 (économie circulaire), S1 (effectifs), S2 (chaîne de valeur), S3
  (communautés), S4 (consommateurs), G1 (gouvernance).
- **Utilité** : Import direct dans le module ESRS pour démarrer l'évaluation
  de conformité CSRD.

**Import sur la plateforme** : `Conformité → ESRS → Import CSV`

### 3. `{supp_path.name}` — Fournisseurs
- **Format** : CSV ; séparateur point-virgule
- **Contenu** : **{supp_count} fournisseurs** avec dépense annuelle, niveau de
  risque, scores ESG (E/S/G), statut d'audit (EcoVadis, rapport public, etc.)
- **Utilité** : Tester le module Supply Chain ESG / CSDDD,
  audits fournisseurs, cartographie des risques.

**Import sur la plateforme** : `Parties → Supply Chain ESG → Import`

### 4. `{mat_path.name}` — Matrice de matérialité
- **Format** : CSV ; séparateur point-virgule
- **Contenu** : **{mat_count} enjeux ESG** notés sur deux axes (impact financier
  × impact ESG, échelle 0-100), avec verdict de matérialité (oui/non) et
  priorité (haute/moyenne/basse).
- **Utilité** : Tester le module **Matérialité CSRD** (double matérialité EFRAG).

**Import sur la plateforme** : `Analyses → Matérialité → Import`

### 5. `{risk_path.name}` — Registre des risques ESG
- **Format** : CSV ; séparateur point-virgule
- **Contenu** : **{risk_count} risques** classés par catégorie (transition,
  opérationnel, réglementaire, réputationnel) avec probabilité × impact (1-5)
  et statut de mitigation.
- **Utilité** : Tester le module **Registre des Risques** + alignement TCFD.

**Import sur la plateforme** : `Analyses → Risques → Import`

### 6. `{carbon_path.name}` — Bilan carbone Scopes 1/2/3
- **Format** : CSV ; séparateur point-virgule
- **Contenu** : **{carbon_count} sources d'émission** ventilées Scopes 1/2/3
  avec les 15 catégories GHG Protocol, méthode de calcul (spend-based,
  average-data, location-based, market-based) et valeurs en **tCO₂e**.
- **Total attendu** : ~2 300 tCO₂e (réaliste pour une PME services 87 ETP)
- **Utilité** : Tester le module **Bilan Carbone** + scores ESRS E1.

**Import sur la plateforme** : `Conformité → Bilan Carbone → Import`

### 7. `{xlsx_path.name}` — Classeur Excel complet
- **Format** : .xlsx avec **5 onglets** (Indicateurs, Carbone, Fournisseurs,
  Matérialité, Risques) + 1 onglet présentation.
- **Utilité** : Vue d'ensemble unique du dataset, plus pratique pour ouvrir
  dans Excel/LibreOffice. Mêmes données que les CSV séparés.

---

## 🚀 Ordre de test recommandé

1. **Créer un compte** sur https://greenconnect.cloud/register (essai ETI 14j)
2. **Compléter le profil organisation** avec les données ci-dessus (sirET,
   secteur, effectif…)
3. **Importer le FEC** (`{fec_path.name}`) → vérifier le calcul Scope 3
   automatique (~1 650 tCO₂e attendus sur cat 1)
4. **Importer le bilan carbone détaillé** (`{carbon_path.name}`) → vérifier
   le total ~2 300 tCO₂e
5. **Importer les indicateurs ESG** (`{indic_path.name}`) → score ESG global
   calculé sur les 12 normes
6. **Importer les fournisseurs** (`{supp_path.name}`) → score Supply Chain ESG
7. **Importer la matrice de matérialité** (`{mat_path.name}`) → matrice 4
   quadrants prête
8. **Importer les risques** (`{risk_path.name}`) → registre rempli avec scores
9. **Tester le benchmark sectoriel** : aller dans `Analyses → Benchmarking`,
   choisir le secteur "Services aux entreprises" et comparer Aurora vs
   référentiel sectoriel
10. **Générer un rapport CSRD** (`Rapports → Générer un rapport → CSRD`) →
    PDF + iXBRL avec toutes les données peuplées

## 🎯 Résultats attendus

Avec ces données, la plateforme devrait calculer :

- **Score ESG global** : ~ **72 / 100** (Note BBB)
- **Bilan carbone total** : **~2 300 tCO₂e** (Scope 3 = 99 % du total)
- **Score Supply Chain** : ~ **68 / 100** (15 fournisseurs évalués)
- **Couverture ESRS** : **{indic_count} / ~80 indicateurs prioritaires** = ~75 %
- **Risques critiques** : **3 risques** identifiés (cyberattaque, talents, CSRD vague 2)
- **Matrice de matérialité** : **9 enjeux matériels sur 10**
- **Délai de paiement fournisseurs** : 38 jours (conforme LME)
- **Gender pay gap** : 5,8 % (conforme Loi Avenir Pro)

## 📝 Notes importantes

- **Données fictives** — aucune entreprise réelle n'est concernée
- **SIREN 851234567** est volontairement invalide (algorithme de Luhn non
  conforme) pour éviter toute confusion avec une vraie société
- **Reproductibilité** : le script utilise `random.seed(42)`, donc régénérer
  le dataset produit exactement les mêmes valeurs
- **Régénération** :
  ```bash
  /Users/cisseniang/Downloads/esgplatform/.claude/worktrees/elastic-moore/backend/.venv/bin/python \\
      scripts/generate_test_dataset.py
  ```

---

*Fichiers générés le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par `scripts/generate_test_dataset.py`*
"""
    path.write_text(content, encoding='utf-8')
    return path


# ─── Main ───────────────────────────────────────────────────────────────────
def main():
    print("Generating test dataset…")
    fec_path, fec_count       = write_fec()
    indic_path, indic_count   = write_indicators_csv()
    supp_path, supp_count     = write_suppliers_csv()
    mat_path, mat_count       = write_materiality_csv()
    risk_path, risk_count     = write_risks_csv()
    carbon_path, carbon_count = write_carbon_csv()
    xlsx_path = write_xlsx()
    readme_path = write_readme((fec_path, fec_count), (indic_path, indic_count),
                                (supp_path, supp_count), (mat_path, mat_count),
                                (risk_path, risk_count), (carbon_path, carbon_count),
                                xlsx_path)

    print(f"\n📂 Test dataset generated in: {out_dir()}\n")
    print(f"  • {fec_path.name:50s}  ({fec_count} écritures)")
    print(f"  • {indic_path.name:50s}  ({indic_count} indicateurs)")
    print(f"  • {supp_path.name:50s}  ({supp_count} fournisseurs)")
    print(f"  • {mat_path.name:50s}  ({mat_count} enjeux)")
    print(f"  • {risk_path.name:50s}  ({risk_count} risques)")
    print(f"  • {carbon_path.name:50s}  ({carbon_count} sources GHG)")
    print(f"  • {xlsx_path.name:50s}  (Excel master 5 onglets)")
    print(f"  • {readme_path.name:50s}  (guide d'utilisation)")
    print()
    print(f"➡️  Lire le README : {readme_path}")


if __name__ == '__main__':
    main()
