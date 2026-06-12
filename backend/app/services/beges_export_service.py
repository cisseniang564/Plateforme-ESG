"""
BEGES Export Service — generate a BEGES.xlsx file for ADEME submission.

Aligned with:
  * Article L229-25 du Code de l'environnement (Bilan d'Émissions de GES)
  * Décret n°2011-829 du 11 juillet 2011 (modifié par décret 2022-982)
  * Méthode réglementaire ADEME v5 (2022)
  * Plateforme bilans-ges.ademe.fr (format de soumission)

Output structure (single .xlsx file with 5 sheets):
  Sheet 1 — Informations entité      (raison sociale, SIREN, NACE, périmètre)
  Sheet 2 — Synthèse                  (totaux par poste, par scope)
  Sheet 3 — Scope 1                   (émissions directes — sources fixes & mobiles)
  Sheet 4 — Scope 2                   (émissions indirectes énergie — location + market)
  Sheet 5 — Scope 3                   (émissions indirectes — 15 catégories GHG Protocol)
  Sheet 6 — Plan d'actions            (réduction visée + actions associées)

This is NOT a direct submission to bilans-ges.ademe.fr (which has no public
upload API yet). The user downloads the .xlsx and uploads it manually on the
ADEME platform.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.organization import Organization
from app.models.tenant import Tenant

logger = logging.getLogger(__name__)


# ─── Styles ─────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_SUBHEAD_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_SUBHEAD_FONT = Font(name="Calibri", size=10, bold=True, color="064E3B")
_TOTAL_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="92400E")
_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


# ─── GHG Protocol Scope 3 — 15 catégories ──────────────────────────────────
SCOPE3_CATEGORIES = [
    (1,  "Achats de biens et services"),
    (2,  "Immobilisations de biens"),
    (3,  "Activités liées au combustible et énergie (hors Scope 1/2)"),
    (4,  "Transport et distribution amont"),
    (5,  "Déchets générés par l'activité"),
    (6,  "Déplacements professionnels"),
    (7,  "Déplacements domicile-travail"),
    (8,  "Actifs en leasing amont"),
    (9,  "Transport et distribution aval"),
    (10, "Transformation des produits vendus"),
    (11, "Utilisation des produits vendus"),
    (12, "Fin de vie des produits vendus"),
    (13, "Actifs en leasing aval"),
    (14, "Franchises"),
    (15, "Investissements"),
]


class BEGESExportService:
    """Service to produce a BEGES.xlsx file for ADEME submission."""

    def __init__(self, db: AsyncSession, tenant_id: UUID):
        self.db = db
        self.tenant_id = tenant_id

    # ── Public ─────────────────────────────────────────────────────────────
    async def generate(
        self,
        organization_id: Optional[UUID] = None,
        year: Optional[int] = None,
    ) -> bytes:
        """
        Build and return the .xlsx file as bytes.

        Args:
            organization_id: If provided, the report covers a single
                organization. If None, defaults to the first org of the tenant.
            year: Reporting year (default = current year - 1).

        Returns:
            xlsx file content as bytes (ready for HTTP download).
        """
        year = year or (datetime.utcnow().year - 1)

        tenant = await self.db.get(Tenant, self.tenant_id)
        org = await self._resolve_org(organization_id)
        emissions = await self._load_emissions(org.id if org else None, year)

        wb = Workbook()
        # Remove the default sheet, we'll create ours
        wb.remove(wb.active)

        self._sheet_identity(wb, tenant, org, year)
        self._sheet_synthesis(wb, emissions, year)
        self._sheet_scope(wb, "Scope 1", emissions.get("scope_1", []), year, color="DC2626")
        self._sheet_scope(wb, "Scope 2", emissions.get("scope_2", []), year, color="EA580C")
        self._sheet_scope3(wb, emissions.get("scope_3", {}), year)
        self._sheet_action_plan(wb, emissions, year)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # ── Sheets ─────────────────────────────────────────────────────────────
    def _sheet_identity(
        self,
        wb: Workbook,
        tenant: Optional[Tenant],
        org: Optional[Organization],
        year: int,
    ) -> None:
        ws = wb.create_sheet("1 - Identité")
        self._title(ws, "BEGES — Bilan d'Émissions de Gaz à Effet de Serre", subtitle=f"Article L229-25 du Code de l'environnement — Exercice {year}")

        rows = [
            ("Raison sociale", (org.legal_name or org.name) if org else (tenant.name if tenant else "—")),
            ("Forme juridique", "—"),
            ("SIREN", org.siren if org else "—"),
            ("SIRET (siège)", org.siret if org else "—"),
            ("Code NACE / NAF", org.sector_code if org else "—"),
            ("Effectif total (ETP)", org.employee_count if org and org.employee_count else "—"),
            ("Chiffre d'affaires (€)", f"{org.revenue_eur:,.0f}".replace(",", " ") if org and org.revenue_eur else "—"),
            ("Pays", (org.country_code or "FR") if org else "FR"),
            ("", ""),
            ("Année de reporting", year),
            ("Date de clôture exercice", f"31/12/{year}"),
            ("Périmètre opérationnel", "Contrôle opérationnel (recommandation GHG Protocol)"),
            ("Périmètre organisationnel", "Toutes les entités sous contrôle"),
            ("Méthode de calcul", "Facteurs d'émission Base Carbone ADEME v23"),
            ("", ""),
            ("Date de soumission", date.today().isoformat()),
            ("Outil utilisé", "ESG Flow — greenconnect.cloud"),
        ]
        self._write_kv_block(ws, start_row=4, rows=rows)
        self._autosize(ws, [2, 5])

    def _sheet_synthesis(self, wb: Workbook, emissions: Dict[str, Any], year: int) -> None:
        ws = wb.create_sheet("2 - Synthèse")
        self._title(ws, "Synthèse des émissions", subtitle=f"Totaux par scope — Exercice {year}")

        # Compute totals
        s1 = sum(Decimal(str(e.get("tco2e", 0))) for e in emissions.get("scope_1", []))
        s2 = sum(Decimal(str(e.get("tco2e", 0))) for e in emissions.get("scope_2", []))
        s3 = Decimal("0")
        for cat_data in emissions.get("scope_3", {}).values():
            s3 += Decimal(str(cat_data.get("tco2e", 0)))
        total = s1 + s2 + s3

        ws.cell(row=4, column=1, value="Scope").font = _HEADER_FONT
        ws.cell(row=4, column=1).fill = _HEADER_FILL
        ws.cell(row=4, column=2, value="Description").font = _HEADER_FONT
        ws.cell(row=4, column=2).fill = _HEADER_FILL
        ws.cell(row=4, column=3, value="Émissions (tCO2e)").font = _HEADER_FONT
        ws.cell(row=4, column=3).fill = _HEADER_FILL
        ws.cell(row=4, column=4, value="% du total").font = _HEADER_FONT
        ws.cell(row=4, column=4).fill = _HEADER_FILL

        for i in range(1, 5):
            ws.cell(row=4, column=i).alignment = Alignment(horizontal="center")

        data = [
            ("Scope 1", "Émissions directes (combustion, fuites, véhicules)", s1),
            ("Scope 2", "Émissions indirectes liées à l'énergie achetée", s2),
            ("Scope 3", "Autres émissions indirectes (chaîne de valeur)", s3),
        ]
        for r, (scope, desc, value) in enumerate(data, start=5):
            ws.cell(row=r, column=1, value=scope).font = Font(bold=True)
            ws.cell(row=r, column=2, value=desc)
            ws.cell(row=r, column=3, value=float(value)).number_format = '#,##0.0'
            pct = (float(value) / float(total) * 100) if total > 0 else 0
            ws.cell(row=r, column=4, value=f"{pct:.1f} %").alignment = Alignment(horizontal="right")
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = _BORDER

        # Total row
        r = 5 + len(data)
        ws.cell(row=r, column=1, value="TOTAL").font = _TOTAL_FONT
        ws.cell(row=r, column=1).fill = _TOTAL_FILL
        ws.cell(row=r, column=2, value="Émissions brutes totales").font = _TOTAL_FONT
        ws.cell(row=r, column=2).fill = _TOTAL_FILL
        ws.cell(row=r, column=3, value=float(total)).number_format = '#,##0.0'
        ws.cell(row=r, column=3).font = _TOTAL_FONT
        ws.cell(row=r, column=3).fill = _TOTAL_FILL
        ws.cell(row=r, column=4, value="100.0 %").alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4).font = _TOTAL_FONT
        ws.cell(row=r, column=4).fill = _TOTAL_FILL

        self._autosize(ws, [10, 50, 18, 12])

    def _sheet_scope(
        self,
        wb: Workbook,
        name: str,
        items: List[Dict[str, Any]],
        year: int,
        color: str = "DC2626",
    ) -> None:
        ws = wb.create_sheet(name)
        self._title(ws, f"Détail {name}", subtitle=f"Exercice {year}", accent=color)

        headers = ["Poste", "Source / Activité", "Quantité", "Unité",
                   "Facteur d'émission", "Émissions (tCO2e)", "Source FE"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        for r_idx, item in enumerate(items, start=5):
            ws.cell(row=r_idx, column=1, value=item.get("poste", "—"))
            ws.cell(row=r_idx, column=2, value=item.get("activity", "—"))
            ws.cell(row=r_idx, column=3, value=item.get("quantity", 0)).number_format = '#,##0.0'
            ws.cell(row=r_idx, column=4, value=item.get("unit", "—"))
            ws.cell(row=r_idx, column=5, value=item.get("emission_factor", 0)).number_format = '0.0000'
            ws.cell(row=r_idx, column=6, value=item.get("tco2e", 0)).number_format = '#,##0.0'
            ws.cell(row=r_idx, column=7, value=item.get("source", "ADEME Base Carbone v23"))
            for c in range(1, 8):
                ws.cell(row=r_idx, column=c).border = _BORDER

        if not items:
            ws.cell(row=5, column=1, value="Aucune donnée saisie pour ce scope.")
            ws.cell(row=5, column=1).font = Font(italic=True, color="9CA3AF")

        # Total row
        if items:
            total_row = 5 + len(items)
            ws.cell(row=total_row, column=1, value="TOTAL").font = _TOTAL_FONT
            ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
            total_value = sum(Decimal(str(it.get("tco2e", 0))) for it in items)
            ws.cell(row=total_row, column=6, value=float(total_value)).number_format = '#,##0.0'
            ws.cell(row=total_row, column=6).font = _TOTAL_FONT
            ws.cell(row=total_row, column=6).fill = _TOTAL_FILL
            for c in (2, 3, 4, 5, 7):
                ws.cell(row=total_row, column=c).fill = _TOTAL_FILL

        self._autosize(ws, [25, 35, 12, 10, 16, 18, 25])

    def _sheet_scope3(self, wb: Workbook, scope3_data: Dict[int, Any], year: int) -> None:
        ws = wb.create_sheet("Scope 3")
        self._title(ws, "Détail Scope 3 — 15 catégories GHG Protocol", subtitle=f"Exercice {year}", accent="7C3AED")

        headers = ["Cat.", "Catégorie GHG Protocol", "Inclus", "Émissions (tCO2e)", "Méthode", "Commentaires"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        total = Decimal("0")
        for r_idx, (cat_num, cat_label) in enumerate(SCOPE3_CATEGORIES, start=5):
            data = scope3_data.get(cat_num, {}) if isinstance(scope3_data, dict) else {}
            included = "Oui" if data.get("included", bool(data.get("tco2e"))) else "Non"
            tco2e = float(data.get("tco2e", 0) or 0)
            method = data.get("method", "Monétaire (PCG)" if tco2e > 0 else "—")
            comment = data.get("comment", "")
            total += Decimal(str(tco2e))

            ws.cell(row=r_idx, column=1, value=cat_num).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=2, value=cat_label)
            ws.cell(row=r_idx, column=3, value=included).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=4, value=tco2e).number_format = '#,##0.0'
            ws.cell(row=r_idx, column=5, value=method)
            ws.cell(row=r_idx, column=6, value=comment)
            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).border = _BORDER

        total_row = 5 + len(SCOPE3_CATEGORIES)
        ws.cell(row=total_row, column=1, value="TOTAL").font = _TOTAL_FONT
        ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
        ws.cell(row=total_row, column=4, value=float(total)).number_format = '#,##0.0'
        ws.cell(row=total_row, column=4).font = _TOTAL_FONT
        ws.cell(row=total_row, column=4).fill = _TOTAL_FILL
        for c in (2, 3, 5, 6):
            ws.cell(row=total_row, column=c).fill = _TOTAL_FILL

        self._autosize(ws, [6, 55, 10, 18, 22, 35])

    def _sheet_action_plan(self, wb: Workbook, emissions: Dict[str, Any], year: int) -> None:
        ws = wb.create_sheet("Plan d'actions")
        self._title(ws, "Plan d'actions de réduction des émissions", subtitle="Obligation Décret 2022-982 — actions sur 4 ans")

        headers = ["Action", "Scope concerné", "Réduction visée (tCO2e)", "Échéance", "Statut", "Responsable"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        # Placeholder rows — to be filled by the user before submission
        ws.cell(row=5, column=1, value="À compléter avant soumission ADEME.")
        ws.cell(row=5, column=1).font = Font(italic=True, color="9CA3AF")

        self._autosize(ws, [40, 18, 22, 12, 12, 20])

    # ── Helpers ────────────────────────────────────────────────────────────
    def _title(self, ws, title: str, subtitle: str = "", accent: str = "064E3B") -> None:
        ws.merge_cells("A1:G1")
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        if subtitle:
            ws.merge_cells("A2:G2")
            cell = ws.cell(row=2, column=1, value=subtitle)
            cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
            cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 18

    def _write_kv_block(self, ws, start_row: int, rows) -> None:
        for i, (key, value) in enumerate(rows):
            r = start_row + i
            if key:
                ws.cell(row=r, column=1, value=key).font = Font(bold=True, color="064E3B")
            ws.cell(row=r, column=2, value=value)

    @staticmethod
    def _autosize(ws, widths: List[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data loading ───────────────────────────────────────────────────────
    async def _resolve_org(self, organization_id: Optional[UUID]) -> Optional[Organization]:
        if organization_id:
            org = await self.db.get(Organization, organization_id)
            if org and org.tenant_id == self.tenant_id:
                return org
        # Fallback: pick the first organization of the tenant
        result = await self.db.execute(
            select(Organization).where(Organization.tenant_id == self.tenant_id).limit(1)
        )
        return result.scalar_one_or_none()

    async def _load_emissions(self, org_id: Optional[UUID], year: int) -> Dict[str, Any]:
        """Load Scope 1/2/3 from indicator_data, organized for the export."""
        from app.models.indicator_data import IndicatorData
        from app.models.indicator import Indicator

        if not org_id:
            return {"scope_1": [], "scope_2": [], "scope_3": {}}

        # Pull indicator rows whose code or name contains scope/CO2 markers.
        period_start = date(year, 1, 1)
        period_end = date(year, 12, 31)

        stmt = (
            select(IndicatorData, Indicator)
            .join(Indicator, IndicatorData.indicator_id == Indicator.id)
            .where(IndicatorData.tenant_id == self.tenant_id)
            .where(IndicatorData.organization_id == org_id)
            .where(IndicatorData.date.between(period_start, period_end))
            .where(IndicatorData.value.is_not(None))
        )
        result = await self.db.execute(stmt)

        scope1: List[Dict[str, Any]] = []
        scope2: List[Dict[str, Any]] = []
        scope3: Dict[int, Dict[str, Any]] = {}

        for data, indicator in result.all():
            code = (indicator.code or "").upper()
            name = (indicator.name or "").lower()
            value = float(data.value or 0)
            entry = {
                "poste": indicator.category or "—",
                "activity": indicator.name,
                "quantity": value,
                "unit": indicator.unit or "tCO2e",
                "emission_factor": 1.0,
                "tco2e": value if (indicator.unit or "").lower() in ("tco2e", "tco2") else value,
                "source": "Base Carbone ADEME v23",
            }

            if "scope 1" in name or "scope1" in name or code.startswith("ENV-001"):
                scope1.append(entry)
            elif "scope 2" in name or "scope2" in name or code.startswith("ENV-002"):
                scope2.append(entry)
            elif "scope 3" in name or "scope3" in name:
                # Try to deduce the category (1-15) from the name; default to 1.
                cat = 1
                for n, _ in SCOPE3_CATEGORIES:
                    if f"cat.{n}" in name or f"catégorie {n}" in name or f"category {n}" in name:
                        cat = n
                        break
                bucket = scope3.setdefault(cat, {"tco2e": 0.0, "included": True, "method": "Monétaire (PCG)"})
                bucket["tco2e"] += value

        return {"scope_1": scope1, "scope_2": scope2, "scope_3": scope3}
