"""
Import templates endpoint — generates pre-filled CSV/Excel templates that
exactly match the auto-detection conventions of ``esg_data_import_service``.

Routes
------
* ``GET /api/v1/import-templates/all-in-one.xlsx`` — multi-sheet Excel workbook
  with one tab per data type (Indicateurs ESRS, GHG Carbone, Matérialité,
  Risques ESG, Fournisseurs). Each sheet ships a header row matching the
  dispatcher's expected columns + 2–4 example rows so the user just has to
  fill the values and re-upload.

The single-type CSV downloads remain in the frontend (client-side blob) for
zero-latency UX; this endpoint is purpose-built for the "tout-en-un" Excel
which non-trivial to construct from JS without a heavy library.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/import-templates", tags=["Import Templates"])


# ─── Sheet specifications ─────────────────────────────────────────────────────
# Each tuple: (sheet_name, header_row, example_rows, hint_text)
# Column names mirror what ``ESGDataImportService._detect_file_intent`` and
# the per-table dispatchers (`_import_materiality`, `_import_risks`,
# `_import_suppliers`) expect — so the resulting workbook can be uploaded
# directly without any column re-mapping.

_SHEET_SPECS: List[Tuple[str, List[str], List[List], str]] = [
    (
        "Indicateurs ESRS",
        ["metric_name", "value_numeric", "unit", "pillar", "category",
         "period_start", "period_end", "data_source"],
        [
            ["Émissions GES Scope 1",          "",   "tCO2e", "environmental", "ghg_scope1",   "2025-01-01", "2025-12-31", ""],
            ["Émissions GES Scope 2",          "",   "tCO2e", "environmental", "ghg_scope2",   "2025-01-01", "2025-12-31", ""],
            ["Consommation énergétique",       "",   "MWh",   "environmental", "energy",       "2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP",             "",   "ETP",   "social",        "workforce",    "2025-01-01", "2025-12-31", ""],
            ["Part de femmes au CODIR",        "",   "%",     "social",        "diversity",    "2025-01-01", "2025-12-31", ""],
            ["Heures de formation par ETP",    "",   "h",     "social",        "training",     "2025-01-01", "2025-12-31", ""],
            ["Membres indépendants du conseil","",   "%",     "governance",    "board",        "2025-01-01", "2025-12-31", ""],
        ],
        "Onglet pour les indicateurs ESG génériques — alimente la page Indicateurs / Scores ESG.",
    ),
    (
        "Bilan Carbone",
        ["scope", "category", "source", "tco2e", "calculation_method", "period"],
        [
            ["scope1", "1.1", "Combustion gaz chauffage bureaux",      26.5, "average-data",  "2025"],
            ["scope1", "1.2", "Carburant flotte (5 véhicules diesel)", 21.0, "average-data",  "2025"],
            ["scope2", "2.1", "Électricité achetée (location-based)",  18.3, "location-based","2025"],
            ["scope3", "3.4", "Voyages d'affaires (avion + train)",    12.7, "distance-based","2025"],
        ],
        "Onglet GHG Protocol Scopes 1/2/3 — facteurs ADEME, alimente la page Indicateurs.",
    ),
    (
        "Matérialité",
        ["name", "category", "financial_impact", "esg_impact", "is_material",
         "priority", "stakeholders", "description"],
        [
            ["Émissions GES et empreinte carbone",        "E1 Climat",        78, 92, "oui", "high",   "salariés,clients,investisseurs", "Émissions Scope 1+2+3"],
            ["Mix énergétique et énergies renouvelables", "E1 Climat",        65, 80, "oui", "high",   "investisseurs,clients",          "Part EnR du mix"],
            ["Gestion responsable des déchets",           "E5 Économie circ.",35, 58, "non", "medium", "salariés",                       "Recyclage et valorisation"],
            ["Attraction et rétention des talents",       "S1 Effectifs",     88, 75, "oui", "high",   "salariés,RH",                    "Turnover et engagement"],
        ],
        "Alimente la Matrice de double matérialité. financial_impact et esg_impact entre 0 et 100.",
    ),
    (
        "Risques ESG",
        ["title", "category", "probability", "impact", "risk_score",
         "severity", "mitigation_status", "owner"],
        [
            ["Hausse prix énergie (Scope 1+2)",   "Transition risk", 4, 3, 12, "élevé",     "en cours", "Direction Achats"],
            ["Pénurie talents tech (turnover)",   "Operational",     4, 4, 16, "critique",  "planifié", "DRH"],
            ["Cyberattaque / fuite données clients","Operational",   3, 5, 15, "critique",  "en cours", "RSSI"],
            ["Non-conformité CSRD vague 2",       "Regulatory",      3, 4, 12, "élevé",     "en cours", "Direction RSE"],
        ],
        "Alimente le Registre des risques. probability et impact sur une échelle 1-5.",
    ),
    (
        "Fournisseurs",
        ["name", "country", "sector", "annual_spend_eur", "risk_level",
         "e_score", "s_score", "g_score", "global_score", "audit_status", "audit_date"],
        [
            ["AMAZON WEB SERVICES", "US", "Cloud / IT",   52000, "medium", 78, 70, 85, 77.7, "self-declared", "2025-06-30"],
            ["EDF",                 "FR", "Énergie",     120000, "low",    82, 78, 88, 82.7, "ecovadis-gold", "2025-04-15"],
            ["TRANSPORTEUR XYZ",    "FR", "Logistique",   38000, "high",   55, 62, 58, 58.3, "rapport-public","2025-03-01"],
        ],
        "Alimente la page Supply Chain ESG. risk_level: low/medium/high/critical.",
    ),
]


# ─── Sector-specific overlays ─────────────────────────────────────────────────
# When ``?sector=<code>`` is provided, we OVERRIDE the example rows of certain
# sheets (typically Indicateurs ESRS + Matérialité + Risques) with rows that
# match the real ESG priorities of that industry. This means a banker doesn't
# get "Scope 1 fioul" pre-filled — they get "Émissions financées portefeuille".

_SECTOR_PROFILES: Dict[str, Dict] = {
    "banking": {
        "label": "Banque & Assurance",
        "esrs_examples": [
            ["Émissions financées (portefeuille de crédit)", "", "tCO2e", "environmental", "ghg_scope3_financed", "2025-01-01", "2025-12-31", ""],
            ["Part des actifs alignés Taxonomie UE",          "", "%",     "environmental", "taxonomy_alignment",  "2025-01-01", "2025-12-31", ""],
            ["Encours financements verts",                    "", "M€",    "environmental", "green_finance",       "2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP",                            "", "ETP",   "social",        "workforce",            "2025-01-01", "2025-12-31", ""],
            ["Part de femmes au CODIR",                       "", "%",     "social",        "diversity",            "2025-01-01", "2025-12-31", ""],
            ["Politique anti-corruption (oui/non)",           "", "bool",  "governance",    "ethics",               "2025-01-01", "2025-12-31", ""],
            ["Membres indépendants du conseil",               "", "%",     "governance",    "board",                "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Émissions financées du portefeuille",       "E1 Climat",          92, 95, "oui", "high",   "investisseurs,régulateur,clients", "Scope 3 catégorie 15"],
            ["Risque climatique des actifs (TCFD)",       "E1 Climat",          85, 88, "oui", "high",   "investisseurs,régulateur",         "Risque physique + transition"],
            ["Inclusion financière",                      "S4 Consommateurs",   60, 75, "oui", "medium", "clients,régulateur",               "Accès aux services bancaires"],
            ["Anti-corruption et conformité",             "G1 Conduite",        78, 90, "oui", "high",   "régulateur,clients",               "ACPR, AMF, sanctions"],
        ],
    },
    "industry": {
        "label": "Industrie / Manufacturing",
        "esrs_examples": [
            ["Émissions GES Scope 1 (combustion)",         "", "tCO2e", "environmental", "ghg_scope1",       "2025-01-01", "2025-12-31", ""],
            ["Émissions GES Scope 2 (électricité)",        "", "tCO2e", "environmental", "ghg_scope2",       "2025-01-01", "2025-12-31", ""],
            ["Consommation énergétique sites",             "", "MWh",   "environmental", "energy",            "2025-01-01", "2025-12-31", ""],
            ["Consommation d'eau process",                 "", "m³",    "environmental", "water",             "2025-01-01", "2025-12-31", ""],
            ["Déchets dangereux générés",                  "", "t",     "environmental", "waste_hazardous",   "2025-01-01", "2025-12-31", ""],
            ["Taux de fréquence accidents du travail",     "", "n/M",   "social",        "safety",            "2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP",                         "", "ETP",   "social",        "workforce",         "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Émissions GES Scope 1/2 sites industriels",  "E1 Climat",          90, 88, "oui", "high",   "régulateur,salariés,riverains",   "Combustion + procédés"],
            ["Consommation et rejets d'eau",               "E3 Eau",             72, 80, "oui", "high",   "régulateur,riverains",            "Stress hydrique zones tendues"],
            ["Santé-sécurité au travail",                  "S1 Effectifs",       85, 92, "oui", "high",   "salariés,CSE",                    "Accidents, MP"],
            ["Économie circulaire et déchets",             "E5 Économie circ.",  68, 75, "oui", "medium", "régulateur,clients",              "DEEE, REP"],
        ],
    },
    "retail": {
        "label": "Retail / Distribution",
        "esrs_examples": [
            ["Émissions Scope 3 transport amont",          "", "tCO2e", "environmental", "ghg_scope3_upstream","2025-01-01", "2025-12-31", ""],
            ["Émissions Scope 3 produits vendus",          "", "tCO2e", "environmental", "ghg_scope3_products","2025-01-01", "2025-12-31", ""],
            ["Part des produits durables (label)",         "", "%",     "environmental", "sustainable_products","2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP (temps plein équivalent)","", "ETP",   "social",        "workforce",         "2025-01-01", "2025-12-31", ""],
            ["Part de CDI dans l'effectif",                "", "%",     "social",        "employment_type",   "2025-01-01", "2025-12-31", ""],
            ["Fournisseurs audités droits humains",        "", "%",     "social",        "human_rights",      "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Émissions de la chaîne logistique (Scope 3)","E1 Climat",        85, 80, "oui", "high",   "régulateur,clients",       "Transport, livraisons"],
            ["Origine et durabilité des produits",         "E5 Économie circ.",72, 88, "oui", "high",   "clients,ONG",              "Bio, équitable, recyclé"],
            ["Conditions de travail des fournisseurs",     "S2 Chaîne valeur", 68, 82, "oui", "high",   "ONG,clients,salariés",     "Sweatshops, sous-traitance"],
            ["Emballages et déchets plastiques",           "E2 Pollution",     65, 78, "oui", "medium", "régulateur,clients",       "Loi AGEC"],
        ],
    },
    "services": {
        "label": "Services / Conseil / IT",
        "esrs_examples": [
            ["Émissions Scope 1 (locaux, véhicules)",      "", "tCO2e", "environmental", "ghg_scope1",       "2025-01-01", "2025-12-31", ""],
            ["Émissions Scope 2 (électricité bureaux)",    "", "tCO2e", "environmental", "ghg_scope2",       "2025-01-01", "2025-12-31", ""],
            ["Voyages d'affaires (avion + train)",         "", "tCO2e", "environmental", "ghg_scope3_travel","2025-01-01", "2025-12-31", ""],
            ["Consommation énergétique cloud (datacenter)","", "MWh",   "environmental", "energy_cloud",     "2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP",                         "", "ETP",   "social",        "workforce",        "2025-01-01", "2025-12-31", ""],
            ["Taux de turnover annuel",                    "", "%",     "social",        "turnover",         "2025-01-01", "2025-12-31", ""],
            ["Heures formation par ETP",                   "", "h",     "social",        "training",         "2025-01-01", "2025-12-31", ""],
            ["Politique anti-corruption (oui/non)",        "", "bool",  "governance",    "ethics",           "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Attraction et rétention des talents",        "S1 Effectifs",     90, 78, "oui", "high",   "salariés,RH",              "Turnover IT > 15 %"],
            ["Empreinte carbone du cloud / IT",            "E1 Climat",        62, 72, "oui", "medium", "clients,investisseurs",    "Datacenters, devices"],
            ["Cybersécurité et données clients",           "G1 Conduite",      85, 88, "oui", "high",   "clients,régulateur",       "RGPD, NIS2"],
            ["Voyages d'affaires (Scope 3)",               "E1 Climat",        55, 65, "non", "medium", "salariés",                 "Politique télétravail"],
        ],
    },
    "energy": {
        "label": "Énergie / Utilities",
        "esrs_examples": [
            ["Émissions Scope 1 production",                "", "tCO2e", "environmental", "ghg_scope1",        "2025-01-01", "2025-12-31", ""],
            ["Intensité carbone du mix énergétique",        "", "gCO2/kWh","environmental","carbon_intensity",  "2025-01-01", "2025-12-31", ""],
            ["Part EnR dans la production totale",          "", "%",     "environmental", "renewable_share",    "2025-01-01", "2025-12-31", ""],
            ["Investissement CapEx vert",                   "", "M€",    "environmental", "green_capex",        "2025-01-01", "2025-12-31", ""],
            ["Consommation d'eau pour refroidissement",     "", "m³",    "environmental", "water_cooling",      "2025-01-01", "2025-12-31", ""],
            ["Taux de fréquence accidents du travail",      "", "n/M",   "social",        "safety",             "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Décarbonation de la production",              "E1 Climat",         95, 92, "oui", "high",   "régulateur,investisseurs", "Trajectoire 1.5 °C"],
            ["Sûreté et fiabilité du réseau",               "G1 Conduite",       88, 80, "oui", "high",   "clients,régulateur",       "ENEDIS, RTE"],
            ["Transition juste pour les salariés",          "S1 Effectifs",      72, 85, "oui", "high",   "salariés,CSE",             "Reconversion charbon→EnR"],
            ["Précarité énergétique des clients",           "S4 Consommateurs",  60, 75, "oui", "medium", "ONG,régulateur",           "Médiateur de l'énergie"],
        ],
    },
    "construction": {
        "label": "Construction / BTP",
        "esrs_examples": [
            ["Émissions Scope 3 matériaux (béton, acier)", "", "tCO2e", "environmental", "ghg_scope3_materials","2025-01-01", "2025-12-31", ""],
            ["Émissions Scope 1 engins et chantiers",      "", "tCO2e", "environmental", "ghg_scope1",        "2025-01-01", "2025-12-31", ""],
            ["Déchets de chantier valorisés",              "", "%",     "environmental", "waste_recovery",    "2025-01-01", "2025-12-31", ""],
            ["Part des projets RE2020 conformes",          "", "%",     "environmental", "re2020_compliance", "2025-01-01", "2025-12-31", ""],
            ["Taux de fréquence accidents du travail",     "", "n/M",   "social",        "safety",            "2025-01-01", "2025-12-31", ""],
            ["Effectif total ETP",                         "", "ETP",   "social",        "workforce",         "2025-01-01", "2025-12-31", ""],
            ["Part de salariés en formation HSE",          "", "%",     "social",        "training",          "2025-01-01", "2025-12-31", ""],
        ],
        "materiality_examples": [
            ["Carbone embarqué des matériaux",             "E1 Climat",         92, 90, "oui", "high",   "régulateur,clients",       "RE2020, ACV"],
            ["Sécurité sur les chantiers",                 "S1 Effectifs",      90, 95, "oui", "high",   "salariés,intérimaires",    "Accidents mortels BTP"],
            ["Conditions de travail sous-traitants",       "S2 Chaîne valeur",  75, 85, "oui", "high",   "ONG,régulateur",           "Travail détaché"],
            ["Économie circulaire des chantiers",          "E5 Économie circ.", 70, 75, "oui", "medium", "régulateur,clients",       "REP PMCB, déchets inertes"],
        ],
    },
}


def _build_workbook(sector: Optional[str] = None) -> bytes:
    """Build the multi-sheet ESG import template as a byte string (xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    # ── Cover sheet ───────────────────────────────────────────────────────
    cover = wb.create_sheet("Mode d'emploi", 0)
    cover.column_dimensions["A"].width = 100
    sector_profile = _SECTOR_PROFILES.get(sector or "") if sector else None
    sector_suffix = f" — secteur {sector_profile['label']}" if sector_profile else ""

    cover["A1"] = f"ESG Flow — Template d'import tout-en-un{sector_suffix}"
    cover["A1"].font = Font(size=18, bold=True, color="0F766E")
    cover["A2"] = (
        f"Généré le {datetime.now().strftime('%d/%m/%Y')} · "
        "https://greenconnect.cloud"
    )
    cover["A2"].font = Font(size=9, color="6B7280", italic=True)

    if sector_profile:
        cover["A3"] = (
            f"Les onglets « Indicateurs ESRS » et « Matérialité » sont pré-remplis "
            f"avec les enjeux typiques du secteur {sector_profile['label'].lower()}."
        )
        cover["A3"].font = Font(size=10, color="0F766E", italic=True, bold=True)

    intro = [
        "",
        "Comment utiliser ce template :",
        "  1. Chaque onglet correspond à un type de données ESG.",
        "  2. Remplissez les colonnes — les exemples sont là pour vous guider.",
        "  3. Sauvegardez en .xlsx ou exportez chaque onglet en .csv (UTF-8).",
        "  4. Uploadez le ou les fichiers via la page Connecteurs → Imports rapides.",
        "",
        "Onglets disponibles :",
    ]
    for i, line in enumerate(intro, start=4):
        cover[f"A{i}"] = line
        cover[f"A{i}"].font = Font(size=11)

    row = 4 + len(intro)
    for name, _hdr, _ex, hint in _SHEET_SPECS:
        cover[f"A{row}"] = f"  • {name} — {hint}"
        cover[f"A{row}"].font = Font(size=10, color="111827")
        row += 1

    cover[f"A{row + 1}"] = (
        "Format : UTF-8 · les colonnes encodées avec accents sont supportées."
    )
    cover[f"A{row + 1}"].font = Font(size=9, color="9CA3AF", italic=True)

    # ── Data sheets ───────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    hint_fill   = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    hint_font   = Font(italic=True, size=9, color="92400E")
    border_thin = Border(*(Side(border_style="thin", color="E5E7EB"),) * 4)

    # Sector-aware overrides for the example rows of certain sheets
    sector_overrides: Dict[str, list] = {}
    if sector_profile:
        sector_overrides["Indicateurs ESRS"] = sector_profile.get("esrs_examples", [])
        sector_overrides["Matérialité"]      = sector_profile.get("materiality_examples", [])

    for name, header, examples, hint in _SHEET_SPECS:
        # Use sector-specific examples if available, else the generic ones
        if name in sector_overrides and sector_overrides[name]:
            examples = sector_overrides[name]
            hint = f"{hint} · Pré-rempli pour le secteur {sector_profile['label']}."

        ws = wb.create_sheet(name)
        # Hint row
        ws.cell(row=1, column=1, value=f"💡 {hint}")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(header))
        ws.cell(row=1, column=1).fill = hint_fill
        ws.cell(row=1, column=1).font = hint_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 22

        # Header row
        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_thin
        ws.row_dimensions[2].height = 22

        # Example rows
        for r_idx, row_data in enumerate(examples, start=3):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border_thin
                cell.alignment = Alignment(vertical="center")

        # Auto-size columns (cap to 50 chars)
        for col_idx, col_name in enumerate(header, start=1):
            max_len = len(str(col_name))
            for ex in examples:
                if c_idx <= len(ex):
                    max_len = max(max_len, len(str(ex[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Freeze header rows
        ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get(
    "/all-in-one.xlsx",
    summary="Template Excel tout-en-un (5 onglets)",
    description=(
        "Génère un classeur Excel avec un onglet par type de données ESG : "
        "Indicateurs ESRS, Bilan Carbone, Matérialité, Risques, Fournisseurs. "
        "Chaque feuille contient les en-têtes attendues par le dispatcher "
        "d'import + des lignes d'exemple. Optionnel: ``?sector=<code>`` "
        "pré-remplit les onglets Indicateurs et Matérialité avec les enjeux "
        "typiques du secteur (banking, industry, retail, services, energy, "
        "construction)."
    ),
    include_in_schema=False,
)
def download_all_in_one_template(
    sector: Optional[str] = Query(
        None,
        description="Secteur d'activité — pré-remplit avec les KPI typiques",
        regex="^(banking|industry|retail|services|energy|construction)$",
    ),
) -> StreamingResponse:
    xlsx_bytes = _build_workbook(sector=sector)
    today = datetime.now().strftime("%Y%m%d")
    suffix = f"-{sector}" if sector else ""
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="esg-flow-template{suffix}-{today}.xlsx"',
            "Cache-Control": "public, max-age=3600",
        },
    )


@router.get(
    "/sectors",
    summary="Liste des profils sectoriels disponibles",
)
def list_sectors() -> Dict[str, list]:
    """Return the list of available sector codes + labels for the UI selector."""
    return {
        "sectors": [
            {"code": code, "label": prof["label"]}
            for code, prof in _SECTOR_PROFILES.items()
        ]
    }
