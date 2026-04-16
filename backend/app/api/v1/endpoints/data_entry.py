from fastapi import APIRouter, Depends, HTTPException, Request, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, extract, func, and_, or_
from typing import List, Optional
from datetime import date, datetime
from uuid import UUID
from pydantic import BaseModel

from app.dependencies import get_db, get_current_user
from app.models.data_entry import DataEntry
from app.models.user import User
from app.services.audit_service import log_change

router = APIRouter()

# ============= SCHEMAS =============

class DataEntryCreate(BaseModel):
    organization_id: Optional[UUID] = None
    period_start: date
    period_end: date
    period_type: str = "annual"
    pillar: str
    category: str
    metric_name: str
    value_numeric: Optional[float] = None
    value_text: Optional[str] = None
    unit: Optional[str] = None
    data_source: Optional[str] = None
    collection_method: Optional[str] = "manual"
    notes: Optional[str] = None

class DataEntryUpdate(BaseModel):
    value_numeric: Optional[float] = None
    value_text: Optional[str] = None
    unit: Optional[str] = None
    notes: Optional[str] = None
    data_source: Optional[str] = None

class DataEntryResponse(BaseModel):
    id: UUID
    organization_id: Optional[UUID]
    period_start: date
    period_end: date
    period_type: str
    pillar: str
    category: str
    metric_name: str
    value_numeric: Optional[float]
    value_text: Optional[str]
    unit: Optional[str]
    data_source: Optional[str]
    collection_method: Optional[str]
    verification_status: str
    notes: Optional[str]
    created_by: Optional[UUID]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class DataEntriesStats(BaseModel):
    total: int
    by_pillar: dict
    by_collection_method: dict
    by_verification_status: dict
    date_range: dict

# ============= ENDPOINTS =============

@router.get("/stats", response_model=DataEntriesStats)
async def get_data_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get statistics about data entries"""
    
    # Total
    total_result = await db.execute(
        select(func.count(DataEntry.id)).where(
            DataEntry.tenant_id == current_user.tenant_id
        )
    )
    total = total_result.scalar() or 0
    
    # By pillar
    pillar_result = await db.execute(
        select(DataEntry.pillar, func.count(DataEntry.id))
        .where(DataEntry.tenant_id == current_user.tenant_id)
        .group_by(DataEntry.pillar)
    )
    by_pillar = {row[0]: row[1] for row in pillar_result}
    
    # By collection method
    method_result = await db.execute(
        select(DataEntry.collection_method, func.count(DataEntry.id))
        .where(DataEntry.tenant_id == current_user.tenant_id)
        .group_by(DataEntry.collection_method)
    )
    by_collection_method = {row[0] or 'unknown': row[1] for row in method_result}
    
    # By verification status
    status_result = await db.execute(
        select(DataEntry.verification_status, func.count(DataEntry.id))
        .where(DataEntry.tenant_id == current_user.tenant_id)
        .group_by(DataEntry.verification_status)
    )
    by_verification_status = {row[0]: row[1] for row in status_result}
    
    # Date range
    date_result = await db.execute(
        select(
            func.min(DataEntry.period_start),
            func.max(DataEntry.period_end)
        ).where(DataEntry.tenant_id == current_user.tenant_id)
    )
    date_row = date_result.first()
    date_range = {
        'min': date_row[0].isoformat() if date_row[0] else None,
        'max': date_row[1].isoformat() if date_row[1] else None
    }
    
    return DataEntriesStats(
        total=total,
        by_pillar=by_pillar,
        by_collection_method=by_collection_method,
        by_verification_status=by_verification_status,
        date_range=date_range
    )

@router.post("/", response_model=DataEntryResponse)
async def create_data_entry(
    entry: DataEntryCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create manual data entry with audit logging"""
    
    if not entry.value_numeric and not entry.value_text:
        raise HTTPException(status_code=400, detail="Provide value_numeric or value_text")
    
    db_entry = DataEntry(
        tenant_id=current_user.tenant_id,
        created_by=current_user.id,
        collection_method=entry.collection_method or "manual",
        **entry.dict(exclude={'collection_method'})
    )
    
    db.add(db_entry)
    await db.commit()
    await db.refresh(db_entry)
    
    await log_change(
        db=db,
        tenant_id=current_user.tenant_id,
        entity_type="data_entries",
        entity_id=db_entry.id,
        action="create",
        user=current_user,
        new_values={
            "metric_name": db_entry.metric_name,
            "value": db_entry.value_numeric or db_entry.value_text,
            "pillar": db_entry.pillar
        },
        ip_address=request.client.host if request.client else None
    )
    
    return db_entry

@router.get("/", response_model=List[DataEntryResponse])
async def get_data_entries(
    pillar: Optional[str] = None,
    category: Optional[str] = None,
    organization_id: Optional[UUID] = None,
    year: Optional[int] = None,
    verification_status: Optional[str] = None,
    collection_method: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List data entries with advanced filters"""
    
    query = select(DataEntry).where(DataEntry.tenant_id == current_user.tenant_id)
    
    if pillar:
        query = query.where(DataEntry.pillar == pillar)
    if category:
        query = query.where(DataEntry.category == category)
    if organization_id:
        query = query.where(DataEntry.organization_id == organization_id)
    if year:
        query = query.where(extract('year', DataEntry.period_start) == year)
    if verification_status:
        query = query.where(DataEntry.verification_status == verification_status)
    if collection_method:
        query = query.where(DataEntry.collection_method == collection_method)
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            or_(
                DataEntry.metric_name.ilike(search_pattern),
                DataEntry.notes.ilike(search_pattern),
                DataEntry.data_source.ilike(search_pattern)
            )
        )
    
    query = query.order_by(DataEntry.created_at.desc())
    query = query.offset(offset).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()

@router.get("/templates/metrics")
async def get_metric_templates():
    """Get metric templates by pillar"""
    return {
        "environmental": {
            "emissions": [
                {"name": "Émissions CO2 Scope 1", "unit": "tCO2e"},
                {"name": "Émissions CO2 Scope 2", "unit": "tCO2e"},
                {"name": "Émissions CO2 Scope 3", "unit": "tCO2e"},
            ],
            "energy": [
                {"name": "Consommation électricité", "unit": "MWh"},
                {"name": "Part énergies renouvelables", "unit": "%"},
            ],
            "water": [
                {"name": "Consommation eau", "unit": "m³"},
            ],
            "waste": [
                {"name": "Déchets produits", "unit": "tonnes"},
                {"name": "Taux de recyclage", "unit": "%"},
            ]
        },
        "social": {
            "workforce": [
                {"name": "Effectif total", "unit": "personnes"},
                {"name": "Turnover", "unit": "%"},
                {"name": "Part de femmes", "unit": "%"},
            ],
            "training": [
                {"name": "Heures de formation", "unit": "heures"},
            ],
            "health_safety": [
                {"name": "Taux de fréquence accidents", "unit": "TF"},
            ]
        },
        "governance": {
            "board": [
                {"name": "Taille du conseil", "unit": "membres"},
                {"name": "Part administrateurs indépendants", "unit": "%"},
            ],
            "ethics": [
                {"name": "Formations éthique", "unit": "personnes"},
            ]
        }
    }

@router.put("/{entry_id}", response_model=DataEntryResponse)
async def update_data_entry(
    entry_id: UUID,
    entry_update: DataEntryUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update data entry with audit logging"""
    result = await db.execute(
        select(DataEntry).where(
            DataEntry.id == entry_id,
            DataEntry.tenant_id == current_user.tenant_id
        )
    )
    db_entry = result.scalar_one_or_none()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Not found")
    
    old_values = {
        "value_numeric": db_entry.value_numeric,
        "value_text": db_entry.value_text,
        "unit": db_entry.unit,
        "notes": db_entry.notes,
        "data_source": db_entry.data_source
    }
    
    update_data = entry_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_entry, field, value)
    
    await db.commit()
    await db.refresh(db_entry)
    
    await log_change(
        db=db,
        tenant_id=current_user.tenant_id,
        entity_type="data_entries",
        entity_id=db_entry.id,
        action="update",
        user=current_user,
        old_values=old_values,
        new_values=update_data,
        ip_address=request.client.host if request.client else None
    )
    
    return db_entry

@router.delete("/{entry_id}")
async def delete_data_entry(
    entry_id: UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete data entry with audit logging"""
    result = await db.execute(
        select(DataEntry).where(
            DataEntry.id == entry_id,
            DataEntry.tenant_id == current_user.tenant_id
        )
    )
    db_entry = result.scalar_one_or_none()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Not found")
    
    await log_change(
        db=db,
        tenant_id=current_user.tenant_id,
        entity_type="data_entries",
        entity_id=db_entry.id,
        action="delete",
        user=current_user,
        old_values={
            "metric_name": db_entry.metric_name,
            "value": db_entry.value_numeric or db_entry.value_text
        },
        ip_address=request.client.host if request.client else None
    )
    
    await db.delete(db_entry)
    await db.commit()
    return {"message": "Deleted"}


# ─── Export endpoint ──────────────────────────────────────────────────────────

@router.get("/export")
async def export_data_entries(
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    pillar: Optional[str] = None,
    year: Optional[int] = None,
    verification_status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export data entries as Excel (.xlsx) or CSV."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    query = select(DataEntry).where(DataEntry.tenant_id == current_user.tenant_id)
    if pillar:
        query = query.where(DataEntry.pillar == pillar)
    if year:
        query = query.where(extract('year', DataEntry.period_start) == year)
    if verification_status:
        query = query.where(DataEntry.verification_status == verification_status)
    query = query.order_by(DataEntry.period_start.desc(), DataEntry.pillar)

    result = await db.execute(query)
    entries = result.scalars().all()

    # Build rows
    columns = [
        "id", "pillar", "category", "metric_name",
        "value_numeric", "value_text", "unit",
        "period_start", "period_end", "period_type",
        "data_source", "collection_method", "verification_status", "notes",
        "created_at",
    ]
    rows = []
    for e in entries:
        rows.append({
            "id": str(e.id),
            "pillar": e.pillar or "",
            "category": e.category or "",
            "metric_name": e.metric_name or "",
            "value_numeric": e.value_numeric,
            "value_text": e.value_text or "",
            "unit": e.unit or "",
            "period_start": str(e.period_start) if e.period_start else "",
            "period_end": str(e.period_end) if e.period_end else "",
            "period_type": e.period_type or "",
            "data_source": e.data_source or "",
            "collection_method": e.collection_method or "",
            "verification_status": e.verification_status or "",
            "notes": e.notes or "",
            "created_at": str(e.created_at)[:19] if e.created_at else "",
        })

    if format == "csv":
        import csv, io
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
        csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel
        return StreamingResponse(
            BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=esg_data_export.csv"},
        )

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ESG Data"

    # Header style
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    HEADER_LABELS = {
        "id": "ID", "pillar": "Pilier", "category": "Catégorie",
        "metric_name": "Indicateur", "value_numeric": "Valeur numérique",
        "value_text": "Valeur texte", "unit": "Unité",
        "period_start": "Début période", "period_end": "Fin période",
        "period_type": "Type période", "data_source": "Source",
        "collection_method": "Méthode", "verification_status": "Statut vérif.",
        "notes": "Notes", "created_at": "Créé le",
    }

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS.get(col, col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Data rows with pillar color coding
    PILLAR_COLORS = {
        "environmental": "e8f5e9",
        "social": "e3f2fd",
        "governance": "f3e5f5",
    }

    for row_idx, row in enumerate(rows, 2):
        pillar_val = row.get("pillar", "")
        row_fill = PatternFill(
            start_color=PILLAR_COLORS.get(pillar_val, "FFFFFF"),
            end_color=PILLAR_COLORS.get(pillar_val, "FFFFFF"),
            fill_type="solid",
        ) if pillar_val in PILLAR_COLORS else None

        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col))
            if row_fill:
                cell.fill = row_fill

    # Freeze header
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=esg_data_export.xlsx"},
    )
